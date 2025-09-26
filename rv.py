from flask import Flask, render_template, redirect, jsonify, url_for, request, flash, session, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from flask_jwt_extended import JWTManager, jwt_required
import os
from PIL import Image
import pytesseract
import random
import requests
import sympy as sp
import re
import cv2
import numpy as np
from datetime import datetime, timedelta
import holidays
import calendar
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import string
import uuid




api_key = 'your_openweathermap_api_key'  # OpenWeatherMap API 키를 여기에 입력하세요.

app = Flask(__name__)

app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Tesseract 경로 설정 (Windows에서만 필요)
pytesseract.pytesseract.tesseract_cmd = r'Tesseract-OCR\tesseract.exe'



app.secret_key = 'your_secret_key'
app.config['JWT_SECRET_KEY'] = 'jwt_secret_key'
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# 사용자 데이터 저장용 (실제 프로젝트에서는 데이터베이스 사용)
users = {}

# 파일 확장자 확인 함수
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# 기본 페이지
@app.route('/')
def index():
    return render_template('index.html', page='home')

# 회원가입
@app.route('/signup', methods=['POST', 'GET'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        email = request.form['email']
        phone = request.form['phone']
        full_name = request.form['full_name']
        if username in users:
            flash('이미 존재하는 아이디입니다!', 'danger')
        else:
            users[username] = {
                'password': generate_password_hash(password),
                'email': email,
                'phone': phone,
                'full_name': full_name
            }
            flash('회원가입이 완료되었습니다!', 'success')
            return redirect(url_for('login'))
    return render_template('index.html', page='signup')

# 로그인
@app.route('/login', methods=['POST', 'GET'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username not in users or not check_password_hash(users[username]['password'], password):
            flash('아이디나 비밀번호가 잘못되었습니다!', 'danger')
        else:
            session['username'] = username
            flash('로그인 성공!', 'success')
            return redirect(url_for('dashboard'))
    return render_template('index.html', page='login')

# 로그아웃
@app.route('/logout')
def logout():
    session.pop('username', None)
    flash('로그아웃 되었습니다.', 'success')
    return redirect(url_for('login'))

# 대시보드 (로그인 후)
@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('index.html', page='dashboard')


@app.route('/delete_account', methods=['POST'])
def delete_account():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    username = session['username']
    # 사용자 삭제
    del users[username]
    session.pop('username', None)  # 로그아웃 처리
    flash('계정이 성공적으로 삭제되었습니다.', 'success')
    return redirect(url_for('signup'))

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

def extract_text_from_image(image_path):
    img = Image.open(image_path)
    
    # 회색조로 변환 (그레이스케일)
    img = img.convert('L')
    
    # 이미지 크기 조정 (해상도가 낮을 때 효과적)
    img = img.resize((img.width * 2, img.height * 2), Image.Resampling.LANCZOS)
    
    # 이진화 (Thresholding)
    np_img = np.array(img)
    _, binary_img = cv2.threshold(np_img, 150, 255, cv2.THRESH_BINARY)
    processed_img = Image.fromarray(binary_img)

    # OCR 설정
    custom_config = r'--oem 3 --psm 6'  # OCR 엔진 모드 및 페이지 분할 모드
    text = pytesseract.image_to_string(processed_img, lang='kor+eng', config=custom_config)
    
    # OCR 후 잘못 인식된 텍스트 후처리
    text = post_process_text(text)
    
    # 문단 나누기
    paragraphs = text.split('\n\n')  # 두 줄로 나누어진 문단 처리
    paragraphs = [p.strip() for p in paragraphs if p.strip()]  # 불필요한 공백 제거
    
    return paragraphs

# 후처리 함수
def post_process_text(text):
    # 잘못된 '0' -> '@' 바꾸기 (예: '0'이 '@'로 인식된 경우)
    text = text.replace('@', '0')
    
    # 문자로 잘못 인식된 숫자들 (예: 'O' -> '0')
    text = text.replace('O', '0')  # 'O'를 '0'으로 변경
    text = text.replace('I', '1')  # 'I'를 '1'으로 변경
    text = text.replace('Z', '2')  # 'Z'를 '2'로 변경
    
    # 필요에 따라 추가적인 후처리 규칙을 추가할 수 있습니다.
    
    return text



# 이미지 업로드 페이지
@app.route('/upload', methods=['GET', "POST"])
def upload_image():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            # 파일 이름 생성 및 저장
            filename = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filename)

            # 이미지에서 텍스트 추출
            extracted_text = extract_text_from_image(filename)

            # 템플릿에 이미지와 텍스트 전달
            return render_template('index.html', page='upload', image_url=filename, extracted_text=extracted_text)
    
    return render_template('index.html', page='upload')


def generate_random_filename(extension):
    random_str = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
    return f"{random_str}.{extension}"

# Cell 클래스 (표 셀을 표현)
class Cell(object):
    def __init__(self):
        self.x = None
        self.y = None
        self.width = None
        self.height = None

        self.central_x = None
        self.central_y = None

        self.text = None
        self.text_height = None
        self.text_align = 'center'
        self.text_valign = 'center'

        self.boundary = {
            'left': False,
            'right': False,
            'upper': False,
            'lower': False
        }

# 이미지 전처리 함수 (해상도 향상 및 노이즈 제거)
def preprocess_image(image_path):
    img = cv2.imread(image_path)
    
    # 해상도 증가: 이미지를 1.2배 크기로 확대 (비율 유지)
    height, width = img.shape[:2]
    new_dim = (int(width * 1.2), int(height * 1.2))
    img = cv2.resize(img, new_dim, interpolation=cv2.INTER_LINEAR)
    
    # 그레이스케일로 변환 (이미지 품질 개선)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # 이진화 (흑백으로 변환하여 글자가 선명하게 보이도록)
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)
    
    # 노이즈 제거 (이미지의 노이즈를 줄여 정확한 텍스트 인식)
    blur = cv2.GaussianBlur(thresh, (5, 5), 0)
    
    return blur, img

# 텍스트 추출 함수
def extract_text(image_path):
    preprocessed_img, img = preprocess_image(image_path)
    
    # Tesseract로 텍스트 추출 (한글 + 영어)
    text = pytesseract.image_to_string(preprocessed_img, lang='kor+eng')
    
    return text, img

# 표가 있을 경우 엑셀로 저장하는 함수
def save_to_excel(text_data, is_table):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Text"
    
    if is_table:
        # 표가 있을 경우 엑셀에 표 형식으로 텍스트 저장
        row_idx = 1
        col_idx = 1
        for text in text_data:
            ws.cell(row=row_idx, column=col_idx, value=text)
            col_idx += 1
            if col_idx > 5:  # 표의 열 수 제한 (예: 5개 열)
                col_idx = 1
                row_idx += 1
    else:
        # 표가 없을 경우 일반 텍스트 저장
        for idx, line in enumerate(text_data):
            ws[f"A{idx + 1}"] = line

    # 테두리 추가 (엑셀 셀에 테두리 그리기)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    # 엑셀 파일 저장
    excel_filename = generate_random_filename("xlsx")
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
    wb.save(output_path)
    return excel_filename

# 불완전한 사각형이나 선이 제대로 잡혀있지 않은 사각형을 완성시켜준다.
def boxing_ambiguous(img, cells):
    for cell in cells:
        # 셀 경계를 완성하는 로직 추가
        pass
    return cells

# 표에 대한 정보를 얻으며, line_image(라인만 있는 이미지)와 erase_line(라인을 제거한 이미지)로 분리한다.
def detect_contours(image_path):
    img = cv2.imread(image_path)
    
    # 그레이스케일 변환
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    # 엣지 검출
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)
    
    # 컨투어(선) 검출
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # 라인만 있는 이미지 생성
    line_image = cv2.drawContours(np.zeros_like(img), contours, -1, (255, 255, 255), 1)
    
    # 선을 지운 이미지 생성
    erase_line_image = cv2.bitwise_not(line_image)
    
    return line_image, erase_line_image

# 텍스트 추출 및 하이라이트 처리 함수
def extract_and_highlight_text(image_path):
    text, img = extract_text(image_path)
    
    # 글자 위치 추출 (박스를 그리기 위해)
    h, w, _ = img.shape
    boxes = pytesseract.image_to_boxes(img)

    # 텍스트 하이라이트 (이미지에 박스 그리기)
    for b in boxes.splitlines():
        b = b.split()
        x, y, w_box, h_box = int(b[1]), int(b[2]), int(b[3]), int(b[4])
        cv2.rectangle(img, (x, h - y), (w_box, h - h_box), (0, 255, 0), 2)  # 글자 박스 그리기
    
    # 하이라이트된 이미지 저장
    highlighted_image_filename = generate_random_filename("jpg")
    highlighted_image_path = os.path.join(app.config['UPLOAD_FOLDER'], highlighted_image_filename)
    cv2.imwrite(highlighted_image_path, img)
    
    return text, highlighted_image_filename


# 하이라이트 페이지
@app.route('/highlight', methods=['GET', 'POST'])
def highlight():
    extracted_text = None
    highlighted_image_path = None
    excel_file_path = None

    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'

        file = request.files['file']
        
        if file.filename == '':
            return 'No selected file'

        if file and allowed_file(file.filename):
            # 파일 저장
            filename = secure_filename(file.filename)
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(image_path)
            
            # 텍스트 추출 및 하이라이트 처리
            text, highlighted_image_filename = extract_and_highlight_text(image_path)
            
            # 텍스트를 한 줄씩 리스트로 저장
            text_data = text.splitlines()

            # 표가 있을 경우 텍스트를 추출하여 엑셀 파일로 저장
            excel_filename = save_to_excel(text_data, is_table=False)  # 표가 없으므로 is_table=False

            # 하이라이트된 이미지 경로
            highlighted_image_path = url_for('uploaded_file', filename=highlighted_image_filename)
            
            # 엑셀 파일 다운로드 경로
            excel_file_path = url_for('download_file', filename=excel_filename)

            extracted_text = '\n'.join(text_data)  # 추출된 텍스트를 하나의 문자열로 합침

    return render_template('index.html', 
                           page='highlight', 
                           extracted_text=extracted_text,
                           highlighted_image_path=highlighted_image_path,
                           excel_file_path=excel_file_path)


# 하이라이트된 이미지 제공
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# 엑셀 파일 다운로드 라우트
@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)




# 날씨 정보를 가져오는 함수 (OpenWeatherMap API 사용)
def get_weather(city_name):
    if not city_name:
        return "도시 이름을 입력해주세요."


    url = f"http://api.openweathermap.org/data/2.5/weather?q={city_name}&appid={api_key}&units=metric&lang=kr"
    
    try:
        response = requests.get(url)
        data = response.json()

        # 날씨 정보가 성공적으로 반환되었는지 확인
        if data['cod'] == 200:
            weather_desc = data['weather'][0]['description']  # 날씨 상태 (맑음, 흐림, 비 등)
            temp = data['main']['temp']  # 기온
            humidity = data['main']['humidity']  # 습도
            wind_speed = data['wind']['speed']  # 바람 속도

            # 유동적인 날씨 응답 생성
            weather_responses = [
                f"{city_name}의 현재 날씨는 {weather_desc}입니다. 기온은 {temp}°C이며, 습도는 {humidity}%입니다. 바람 속도는 {wind_speed}m/s입니다.",
                f"지금 {city_name}의 날씨는 {weather_desc}입니다. 기온은 {temp}°C로 예상되며, 습도는 {humidity}%입니다. 바람 속도는 {wind_speed}m/s입니다.",
                f"{city_name}의 날씨는 현재 {weather_desc}입니다. 기온은 {temp}°C이며, 습도는 {humidity}%입니다. 바람 속도는 {wind_speed}m/s입니다."
            ]
            return random.choice(weather_responses)  # 랜덤으로 유동적인 응답 선택
        else:
            return get_default_weather_forecast(city_name)
    except requests.exceptions.RequestException as e:
        print(f"Error fetching weather: {e}")
        return get_default_weather_forecast(city_name)
    

# 대체 날씨 예보를 제공하는 함수
def get_default_weather_forecast(city_name):
    # 랜덤한 날씨 예보를 제공
    weather_conditions = ['맑음', '구름 많음', '비', '눈', '폭풍']
    temperature = random.randint(-5, 35)
    humidity = random.randint(50, 80)
    wind_speed = random.randint(1, 15)

    weather_forecast = f"{city_name}의 예보: 오늘은 {random.choice(weather_conditions)} 예상됩니다. "
    weather_forecast += f"기온은 약 {temperature}°C이며, 습도는 {humidity}%입니다. "
    weather_forecast += f"바람 속도는 {wind_speed}m/s로 예상됩니다."
    
    return weather_forecast


def get_response(user_input, city_name, api_key):
    if user_input == "오늘 날씨 어때요?":
        return get_weather(city_name, api_key)
    elif user_input in responses:
        return responses[user_input]
    else:
        return "죄송합니다, 그질문에 대한 답변을 알지 못합니다."

responses = {
    "안녕하세요": "안녕하세요! 무엇을 도와드릴까요?",
    "오늘 날씨 어때요?": "오늘 날씨는 맑고 기온은 약간 따뜻합니다.",
    "AI란 무엇인가요?": "AI는 '인공지능'의 약자로, 인간의 지능을 모방하는 기술입니다.",
    "감사합니다": "별말씀을요, 언제든지 도와드리겠습니다!",
    "사용법": "이 시스템은 두 가지 주요 기능을 제공합니다: 1) 이미지에서 텍스트 추출, 2) 챗봇을 통한 질의응답. 이미지를 업로드하여 텍스트를 추출하고, 질문을 통해 답변을 받을 수 있습니다.",
    "기능": "이 시스템의 주요 기능은 이미지에서 텍스트를 추출하고, 다양한 질문에 답변을 제공하는 챗봇 기능입니다.",
    "도움말": "사용법이나 기능에 관한 질문을 하시면 도움이 될 수 있는 정보를 드릴 수 있습니다. 예: '사용법' 또는 '기능'을 입력해주세요.",
    "기본": "이 챗봇은 기본적인 질문 응답 기능을 제공합니다. 예를 들어, 사용법이나 기능에 대한 질문을 할 수 있습니다.",
    "이미지 업로드 방법": "이미지 업로드 방법은 간단합니다. 화면에서 '이미지 업로드' 섹션에 파일을 선택한 후 '이미지 업로드' 버튼을 클릭하세요. 이미지는 자동으로 서버로 전송되어 텍스트가 추출됩니다.",
    "챗봇 질의응답 방법": "챗봇에 질문을 하려면, 화면 하단의 텍스트 입력란에 질문을 입력하고 '응답하기' 버튼을 클릭하세요. Enter 키를 눌러서도 질문을 제출할 수 있습니다.",
    # 계산기 관련 응답 추가
    "계산기 사용법": "계산기 기능을 사용하려면 수식을 입력해 주세요. 예: '2+3', '5*5', 'x^2+2x+1' 등의 계산식을 입력하면 결과를 알려드립니다.",
    "계산기": "수식을 입력하시면 계산 결과를 알려드리겠습니다. 예: '2+3', '10-5', 'sqrt(16)'",
    "계산": "계산하고자 하는 수식을 입력해 주세요. 예: '2*3', '100/4', 'sin(30)'",
    # 응답 불가한 경우를 위한 기본 응답 추가
    "응답 불가": "죄송합니다, 응답할 수 없습니다. 다른 질문을 해주세요.",
    "캘린더 사용법": "이전 버튼을 누르면 지난달 화면으로 넘어가고 다음 버튼을 누르면 다음달 화면으로 넘어갑니다 그리고 해당일자를 클릭해서 일정내용을 기록한 뒤 저장버튼을 누르면 해당일정이 창에 표기가 됩니다 삭제도 가능합니다"
}





# 수학식 계산기 기능을 처리하는 함수
def calculate_expression(query):
    try:
        # 수식에 대한 계산 처리
        # 수식에서 숫자와 연산자만 추출하여 계산하도록 개선
        expression = re.sub(r'[^0-9\+\-\*/\(\)\.\^]', '', query)  # 숫자와 연산자만 남기기
        result = sp.sympify(expression)  # 수식을 안전하게 처리
        result = result.evalf()  # 결과 계산
        return f"계산 결과는 {result}입니다."
    except Exception as e:
        return f"계산 중 오류가 발생했습니다: {e}"




# 키워드 기반 응답 함수
def get_keyword_response(query):
    query = query.lower()  # 사용자의 질문을 소문자로 변환해서 비교
    
    # 업로드와 관련된 키워드 확인
    if "업로드" in query or "파일" in query or "이미지" in query or "사진" in query:
        return responses["이미지 업로드 방법"]
    
    # 계산기와 관련된 키워드 확인
    if "계산기" in query or "계산" in query:
        return responses["계산기"]
    
    if "캘린더" in query or "일정" in query:
        return responses["캘린더 사용법"]
    
    # 날씨 관련 질문
    if "날씨" in query or "기온" in query or "날씨 어때?" in query:
        return responses["오늘 날씨 어때요?"]
    
    # 기존 키워드 기반 응답
    if "ai" in query or "인공지능" in query:
        return responses["AI란 무엇인가요?"]
    elif "사용법" in query:
        return responses["사용법"]
    elif "기능" in query:
        return responses["기능"]
    elif "감사" in query:
        return responses["감사합니다"]
    elif "안녕" in query or "여보세요" in query:
        return responses["안녕하세요"]
    else:
        return "죄송합니다, 해당 질문에 대한 답변을 찾을 수 없습니다. 다른 질문을 해주세요."

# 챗봇 응답 처리 함수
def chat_response(query):
    # 계산기 관련 요청 처리
    if re.search(r'[\d\+\-\*/\(\)\^\.]+', query):  # 숫자 및 연산자가 포함된 경우
        return calculate_expression(query)
    
    if "날씨" in query or "기온" in query:
        # 정규식으로 도시명 추출 (한글 단어 중에서 "날씨"나 "기온"이 포함된 부분)
        match = re.search(r"(대구|제주도|울릉도|강원도|서울|부산|인천|광주|대전|울산|수원|창원|고양|용인|시흥|평택|전주|청주|천안|포항|남양주|김해|광명|전주|구미|등)", query)
        if match:
            city_name = match.group(0)  # 추출된 도시명
            return get_weather(city_name)
        else:
            return "도시명을 명확히 입력해주세요. 예: '대구 날씨 알려줘'"
    else:
        return get_keyword_response(query)
    

@app.route('/chatbot-api', methods=['POST'])
def chatbot_api():
    user_query = request.json.get('query')
    if not user_query:
        return jsonify({'error': '질문을 입력하세요'}), 400

    response_text = chat_response(user_query)
    return jsonify({'response': response_text})    


# 챗봇 페이지
@app.route('/chatbot')
def chatbot():

    return render_template('index.html', page='chatbot')


holidays = {
    '01-01': '새해 첫날',
    '03-01': '삼일절',
    '05-05': '어린이날',
    '06-06': '현충일',
    '08-15': '광복절',
    '10-03': '개천절',
    '10-09': '한글날',
    '12-25': '크리스마스'
}

# 이벤트 저장소 (딕셔너리 형태)
events = {}



def generate_calendar(year, month):
    # 캘린더 생성
    cal = calendar.Calendar(firstweekday=6)  # 일요일부터 시작
    days = cal.monthdayscalendar(year, month)
    
    # 각 날짜에 주말 및 공휴일 여부 추가
    result = []
    for week in days:
        week_data = []
        for day in week:
            if day != 0:
                date_str = f"{month:02d}-{day:02d}"  # 월-일 형식으로 날짜 문자열 생성
                is_weekend = calendar.weekday(year, month, day) >= 5  # 5, 6은 토요일, 일요일
                is_holiday = date_str in holidays
                holiday_name = holidays.get(date_str)
                
                # 해당 날짜의 이벤트가 있으면 가져오기
                event = events.get(f"{year}-{month:02d}-{day:02d}", '')  # event_dates 형태로 키를 가져옴

                week_data.append({
                    'day': day,
                    'is_weekend': is_weekend,
                    'is_holiday': is_holiday,
                    'holiday_name': holiday_name,
                    'event': event  # 이벤트 추가
                })
            else:
                week_data.append(None)
        result.append(week_data)
    return result


# 캘린더 페이지
@app.route('/calendar', methods=['GET'])
def calendar_view():
    # 현재 날짜를 가져오기
    now = datetime.now()

    # 쿼리 파라미터에서 연도와 월을 가져오되 없으면 현재 연도와 월을 사용
    year = int(request.args.get('year', now.year))
    month = int(request.args.get('month', now.month))
    
    # 달력 생성
    days = generate_calendar(year, month)

    return render_template('index.html', page='calendar', year=year, month=month, days=days, events=events)

@app.route('/save_event', methods=['POST'])
def save_event():
    data = request.get_json()
    event_date = data['date']
    event_title = data['title']

    event_id = str(uuid.uuid4())  # 고유 ID 생성
    events[event_id] = {'date': event_date, 'title': event_title}

    # 일정 저장
    if event_date in events:
        events[event_date].append(event_title)
    else:
        events[event_date] = [event_title]

    return jsonify(success=True)

@app.route('/update_event', methods=['POST'])
def update_event():
    data = request.get_json()
    event_date = data['date']
    event_title = data['title']

    # 일정 수정
    if event_date in events:
        events[event_date] = [event_title]  # 기존 일정 수정
        return jsonify(success=True)
    return jsonify(success=False)

@app.route('/delete_event', methods=['POST'])
def delete_event():
    data = request.get_json()
    event_date = data['date']

    # 일정 삭제
    if event_date in events:
        del events[event_date]
        return jsonify(success=True)
    return jsonify(success=False)


if __name__ == '__main__':
    app.run(debug=True)
